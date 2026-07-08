#!/usr/bin/env python3
"""
Sync townstone upgrades between cfg and xlsx.

Default paths:
- cfg:  pkg/opt/townstones/upgrades.cfg
- xlsx: pkg/opt/townstones/upgrades.xlsx

Examples:
  python pythonscripts/_sync_townstone_upgrades.py --direction cfg-to-xlsx
  python pythonscripts/_sync_townstone_upgrades.py --direction xlsx-to-cfg
  python pythonscripts/_sync_townstone_upgrades.py --direction auto
"""

from __future__ import annotations

import argparse
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill


@dataclass
class UpgradeBlock:
    upgrade_id: str
    fields: Dict[str, str] = field(default_factory=dict)


UPGRADE_START_RE = re.compile(r"^\s*Upgrade\s+(\S+)\s*$", re.IGNORECASE)


def repo_root() -> Path:
    return Path(__file__).resolve().parent.parent


def default_cfg_path() -> Path:
    return repo_root() / "pkg" / "opt" / "townstones" / "upgrades.cfg"


def default_xlsx_path() -> Path:
    return repo_root() / "pkg" / "opt" / "townstones" / "upgrades.xlsx"


def parse_cfg(path: Path) -> tuple[list[str], list[UpgradeBlock], list[str]]:
    lines = path.read_text(encoding="utf-8", errors="replace").splitlines()

    header_lines: list[str] = []
    upgrades: list[UpgradeBlock] = []
    field_order: list[str] = []
    in_first_block = False

    i = 0
    while i < len(lines):
        line = lines[i]
        m = UPGRADE_START_RE.match(line)
        if not m:
            if not in_first_block:
                header_lines.append(line)
            i += 1
            continue

        in_first_block = True
        upgrade_id = m.group(1)
        i += 1

        while i < len(lines) and not lines[i].strip():
            i += 1

        if i >= len(lines) or lines[i].strip() != "{":
            # Malformed block; skip safely.
            continue

        i += 1
        fields: dict[str, str] = {}

        while i < len(lines):
            raw = lines[i]
            stripped = raw.strip()

            if stripped == "}":
                i += 1
                break

            if not stripped or stripped.startswith("#"):
                i += 1
                continue

            parts = stripped.split(None, 1)
            if len(parts) == 2:
                key, value = parts
                fields[key] = value.strip()
                if key not in field_order:
                    field_order.append(key)

            i += 1

        upgrades.append(UpgradeBlock(upgrade_id=upgrade_id, fields=fields))

    if not field_order:
        field_order = ["Name", "Desc", "Cost", "Type", "Graphic"]
    elif "Type" not in field_order:
        field_order.append("Type")

    return header_lines, upgrades, field_order


def write_cfg(path: Path, header_lines: list[str], upgrades: list[UpgradeBlock], field_order: list[str]) -> None:
    out: list[str] = []

    if header_lines:
        out.extend(header_lines)
        if out and out[-1].strip():
            out.append("")

    for idx, upg in enumerate(upgrades):
        out.append(f"Upgrade {upg.upgrade_id}")
        out.append("{")

        for key in field_order:
            value = upg.fields.get(key, "").strip()
            if value:
                out.append(f"    {key}    {value}")

        out.append("}")
        if idx != len(upgrades) - 1:
            out.append("")

    path.write_text("\n".join(out) + "\n", encoding="utf-8")


def write_xlsx(path: Path, header_lines: list[str], upgrades: list[UpgradeBlock], field_order: list[str], cfg_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Upgrades"

    headers = ["Upgrade"] + field_order
    ws.append(headers)

    for upg in upgrades:
        row = [upg.upgrade_id] + [upg.fields.get(k, "") for k in field_order]
        ws.append(row)

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"

    for col in ws.columns:
        max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 14), 60)

    meta = wb.create_sheet("Metadata")
    meta["A1"] = "SourceCfg"
    meta["B1"] = str(cfg_path)
    meta["A2"] = "HeaderComments"
    meta["B2"] = "\n".join(header_lines)
    meta["A3"] = "Usage"
    meta["B3"] = (
        "Edit rows in Upgrades sheet. Keep header names unchanged. "
        "Run _sync_townstone_upgrades.py --direction xlsx-to-cfg to push back to cfg."
    )

    for ref in ("A1", "A2", "A3"):
        meta[ref].font = Font(bold=True)

    meta.column_dimensions["A"].width = 20
    meta.column_dimensions["B"].width = 120
    meta["B2"].alignment = Alignment(wrap_text=True, vertical="top")
    meta.row_dimensions[2].height = 90

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def read_xlsx(path: Path) -> tuple[list[str], list[UpgradeBlock], list[str]]:
    wb = openpyxl.load_workbook(path)
    if "Upgrades" not in wb.sheetnames:
        raise ValueError("Workbook is missing required 'Upgrades' sheet.")

    ws = wb["Upgrades"]
    if ws.max_row < 1:
        raise ValueError("Upgrades sheet is empty.")

    headers: list[str] = []
    for cell in ws[1]:
        val = "" if cell.value is None else str(cell.value).strip()
        headers.append(val)

    if not headers or headers[0] != "Upgrade":
        raise ValueError("Column A header must be exactly 'Upgrade'.")

    field_order = [h for h in headers[1:] if h]
    if not field_order:
        raise ValueError("No field columns found after 'Upgrade'.")

    upgrades: list[UpgradeBlock] = []
    for r in range(2, ws.max_row + 1):
        upgrade_id = ws.cell(row=r, column=1).value
        if upgrade_id is None or not str(upgrade_id).strip():
            continue

        fields: dict[str, str] = {}
        for c, key in enumerate(field_order, start=2):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            value = str(v).strip()
            if value:
                fields[key] = value

        upgrades.append(UpgradeBlock(upgrade_id=str(upgrade_id).strip(), fields=fields))

    header_lines: list[str] = []
    if "Metadata" in wb.sheetnames:
        meta = wb["Metadata"]
        if str(meta["A2"].value).strip() == "HeaderComments":
            raw = meta["B2"].value
            if raw:
                header_lines = str(raw).splitlines()

    if not header_lines:
        header_lines = [
            "# Townstone upgrades",
            "# Generated by pythonscripts/_sync_townstone_upgrades.py",
        ]

    return header_lines, upgrades, field_order


def choose_auto_direction(cfg_path: Path, xlsx_path: Path) -> str:
    if not cfg_path.exists() and not xlsx_path.exists():
        raise FileNotFoundError("Neither cfg nor xlsx exists.")
    if cfg_path.exists() and not xlsx_path.exists():
        return "cfg-to-xlsx"
    if xlsx_path.exists() and not cfg_path.exists():
        return "xlsx-to-cfg"

    cfg_mtime = cfg_path.stat().st_mtime
    xlsx_mtime = xlsx_path.stat().st_mtime
    if cfg_mtime >= xlsx_mtime:
        return "cfg-to-xlsx"
    return "xlsx-to-cfg"


def main() -> int:
    parser = argparse.ArgumentParser(description="Sync townstone upgrades cfg <-> xlsx")
    parser.add_argument(
        "--direction",
        choices=["cfg-to-xlsx", "xlsx-to-cfg", "auto"],
        default="auto",
        help="Sync direction (auto picks newest file)",
    )
    parser.add_argument("--cfg", default=str(default_cfg_path()), help="Path to upgrades.cfg")
    parser.add_argument("--xlsx", default=str(default_xlsx_path()), help="Path to upgrades.xlsx")
    parser.add_argument("--dry-run", action="store_true", help="Parse/validate and report without writing")

    args = parser.parse_args()

    cfg_path = Path(args.cfg)
    xlsx_path = Path(args.xlsx)

    direction = args.direction
    if direction == "auto":
        direction = choose_auto_direction(cfg_path, xlsx_path)

    if direction == "cfg-to-xlsx":
        if not cfg_path.exists():
            raise FileNotFoundError(f"Missing cfg file: {cfg_path}")
        header_lines, upgrades, field_order = parse_cfg(cfg_path)
        if args.dry_run:
            print(f"DRY RUN: would write xlsx from cfg ({len(upgrades)} upgrades)")
            return 0
        write_xlsx(xlsx_path, header_lines, upgrades, field_order, cfg_path)
        print(f"Wrote xlsx: {xlsx_path}")
        print(f"Upgrades: {len(upgrades)}")
        return 0

    if direction == "xlsx-to-cfg":
        if not xlsx_path.exists():
            raise FileNotFoundError(f"Missing xlsx file: {xlsx_path}")
        header_lines, upgrades, field_order = read_xlsx(xlsx_path)
        if args.dry_run:
            print(f"DRY RUN: would write cfg from xlsx ({len(upgrades)} upgrades)")
            return 0
        write_cfg(cfg_path, header_lines, upgrades, field_order)
        print(f"Wrote cfg: {cfg_path}")
        print(f"Upgrades: {len(upgrades)}")
        return 0

    raise ValueError(f"Unsupported direction: {direction}")


if __name__ == "__main__":
    raise SystemExit(main())
